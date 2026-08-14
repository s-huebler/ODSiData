"""
Repair Merging/data_dictionaries.xlsx:
  - Drop dangling drawing/vmlDrawing/printerSettings relationships from sheetN.xml.rels
  - Remove stale <dimension ref="A1"/> from each sheetN.xml
  - Export existing meaning/notes annotations to annotations_backup.csv
  - Verify with openpyxl (no read_only) and expected row counts
"""
import shutil, zipfile, re, io, csv, sys
from pathlib import Path

INPUT   = Path("Merging/data_dictionaries.xlsx")
BACKUP  = Path("Merging/data_dictionaries_corrupt_backup.xlsx")
ANN_CSV = Path("Merging/annotations_backup.csv")

EXPECTED = {
    "Artacho": 26, "DAmico": 56, "Fujimoto": 22,
    "Ingham": 67, "Liu": 46, "Vallet": 42,
}

BAD_TARGET_PATTERNS = re.compile(
    r'Target="\.\./(drawings|printerSettings)/|Target="\.\./(drawings)/.*vml',
    re.IGNORECASE
)

def strip_bad_rels(xml_bytes: bytes) -> bytes:
    """Remove Relationship elements whose Target points at drawings/ or printerSettings/."""
    text = xml_bytes.decode("utf-8")
    # Remove individual <Relationship .../> lines whose Target is bad
    cleaned = re.sub(
        r'<Relationship\s[^>]*Target="\.\./(?:drawings|printerSettings)/[^"]*"[^/]*/?>',
        "",
        text,
    )
    return cleaned.encode("utf-8")

def strip_dimension(xml_bytes: bytes) -> bytes:
    """Remove <dimension ref="..."/> from sheet XML."""
    text = xml_bytes.decode("utf-8")
    cleaned = re.sub(r'<dimension[^/]*/>', "", text)
    return cleaned.encode("utf-8")

# ── Step 1: backup ─────────────────────────────────────────────────────────────
print(f"Copying {INPUT} → {BACKUP}")
shutil.copy2(INPUT, BACKUP)

# ── Step 2: rewrite zip ────────────────────────────────────────────────────────
buf = io.BytesIO()
with zipfile.ZipFile(INPUT, "r") as zin, zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
    for item in zin.infolist():
        data = zin.read(item.filename)

        if re.match(r"xl/worksheets/_rels/sheet\d+\.xml\.rels", item.filename):
            data = strip_bad_rels(data)
            print(f"  stripped rels: {item.filename}")

        elif re.match(r"xl/worksheets/sheet\d+\.xml", item.filename):
            data = strip_dimension(data)
            print(f"  stripped dimension: {item.filename}")

        zout.writestr(item, data)

INPUT.write_bytes(buf.getvalue())
print(f"Rewrote {INPUT}")

# ── Step 3: verify with openpyxl ───────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("openpyxl not installed — skipping openpyxl verification")
    openpyxl = None

if openpyxl:
    print("\n=== openpyxl verification ===")
    wb = openpyxl.load_workbook(INPUT)           # must NOT use read_only
    ok = True
    for sheet_name, expected_rows in EXPECTED.items():
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = len(rows) - 1               # subtract header
        status = "OK" if data_rows == expected_rows else f"FAIL (got {data_rows})"
        print(f"  {sheet_name:<12} {data_rows} rows  [{status}]")
        if data_rows != expected_rows:
            ok = False
    if ok:
        print("  All row counts match.")
    else:
        print("  ROW COUNT MISMATCH — aborting before annotation export")
        sys.exit(1)

# ── Step 4: export annotations_backup.csv ─────────────────────────────────────
print(f"\n=== Exporting annotations to {ANN_CSV} ===")

# Use openpyxl if available; otherwise fall back to a raw XML read.
if openpyxl:
    wb = openpyxl.load_workbook(INPUT)
    rows_written = 0
    with open(ANN_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["study","column_name","meaning","notes"])
        writer.writeheader()
        for sheet_name in EXPECTED:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            header = rows[0]
            col_idx  = {h: i for i, h in enumerate(header)}
            for row in rows[1:]:
                meaning = row[col_idx["meaning"]]
                notes   = row[col_idx["notes"]]
                if (meaning and str(meaning).strip()) or (notes and str(notes).strip()):
                    writer.writerow({
                        "study":       sheet_name,
                        "column_name": row[col_idx["column_name"]],
                        "meaning":     str(meaning).strip() if meaning else "",
                        "notes":       str(notes).strip()   if notes   else "",
                    })
                    rows_written += 1
    print(f"  Wrote {rows_written} annotated rows to {ANN_CSV}")
else:
    print("  openpyxl unavailable — cannot export annotations")

print("\nDone.")
