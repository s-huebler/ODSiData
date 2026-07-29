import re
import random
import pandas as pd
from openpyxl import load_workbook, Workbook

INPUT = "/Users/sophiehuebler/Desktop/References_Pubmed2.xlsx"
OUTPUT = "/Users/sophiehuebler/Desktop/References_Pubmed2_parsed.xlsx"

# ── helpers ──────────────────────────────────────────────────────────────────

def camel_case_name(name):
    """'van Den Brink' → 'vanDenBrink', 'Stein-Thoeringer' → 'SteinThoeringer'"""
    parts = re.split(r'[\s\-]+', name.strip())
    return ''.join(p for p in parts if p)

def extract_year(text):
    m = re.search(r'\b(1[89]\d\d|20[0-2]\d)\b', text)
    return m.group(1) if m else None

# ── Format B parser (Ji_2024: "I. Lastname, I2. Lastname2, ...")  ─────────────

def is_format_b(c):
    return bool(re.match(r'^[A-Z](?:\.[A-Z])?\.\s+[A-Z][a-z]', c))

def parse_format_b(c):
    """Returns (first_lastname, author_count, has_et_al)."""
    pos = 0
    authors = []
    has_et_al = False

    while pos < len(c):
        m = re.match(r'([A-Z](?:\.[A-Z])?\.)\s+([A-Z][a-z]+)', c[pos:])
        if not m:
            break
        # Take only the first word of the matched lastname
        authors.append(m.group(2).split()[0])
        pos += m.end()
        sep = re.match(r',\s*', c[pos:])
        if sep:
            pos += sep.end()
            et = re.match(r'et\.?\s*al\.?', c[pos:], re.IGNORECASE)
            if et:
                has_et_al = True
                break
        else:
            break

    n = len(authors)
    if has_et_al or n >= 3:
        return (authors[0] if authors else None), '3+', has_et_al
    return (authors[0] if authors else None), n, False

# ── Format A parser (most sheets: "Lastname Initials, ...")  ─────────────────

# Initials: starts with a capital, may be followed by caps, dots, hyphens
INIT_RE = r'[A-Z][A-Z\.\-]*'

def _extract_lastname_from_start(c):
    """Find 'Lastname Initials<sep>' at the very start of c.

    The lookahead (?=\\s|[,.]) after initials prevents matching a capital letter
    that is the start of a multi-character word (e.g. 'B' in 'Bekkum').
    """
    # Initials must be followed by whitespace, comma, period, or end of string
    INIT_SEP = r'[A-Z][A-Z\.\-]*(?=[\s,\.]|$)'
    m = re.match(
        r'^([\w\'\u00C0-\u024F]+(?:[\s\-][\w\'\u00C0-\u024F]+)*?)'
        r'\s+(' + INIT_SEP + r')'
        r'(?:\s*(?:Jr\.?|Sr\.?|III|IV))?\s*[,\.\s]',
        c, re.UNICODE
    )
    if m:
        return m.group(1).strip()
    # Looser fallback: same lookahead but no trailing separator required
    m2 = re.match(
        r'^([\w\'\u00C0-\u024F]+(?:[\s\-][\w\'\u00C0-\u024F]+)*?)'
        r'\s+(' + INIT_SEP + r')',
        c, re.UNICODE
    )
    if m2:
        return m2.group(1).strip()
    # Absolute fallback: first word
    w = re.match(r'^([\w\'\u00C0-\u024F\-]+)', c, re.UNICODE)
    return w.group(1) if w else None

def _token_looks_like_author(tok):
    """Does 'tok' end with initials? (covers non-first author tokens)"""
    return bool(re.search(r'\s+' + INIT_RE + r'\s*\.?\s*$', tok, re.UNICODE))

def parse_format_a(c):
    """Returns (first_lastname, author_count, has_et_al)."""
    c = re.sub(r'^\d+\.?\s*', '', c.strip())  # strip leading "1."
    has_et_al = bool(re.search(r'\bet\.?\s*al\b', c[:300], re.IGNORECASE))

    tokens = re.split(r',\s*', c)
    author_count = 0
    first_lastname = None

    for i, tok in enumerate(tokens):
        tok = tok.strip()
        if not tok:
            continue
        if re.search(r'\bet\.?\s*al\b', tok, re.IGNORECASE):
            has_et_al = True
            break

        if i == 0:
            ln = _extract_lastname_from_start(tok)
            if ln:
                first_lastname = ln
                author_count += 1
            else:
                author_count += 1  # count but no name
        else:
            if _token_looks_like_author(tok):
                author_count += 1
            else:
                # Token may be "Lastname Initials. Title..." — check from start
                INIT_SEP = r'[A-Z][A-Z\.\-]*(?=[\s,\.]|$)'
                m_start = re.match(
                    r'^([\w\'\u00C0-\u024F]+(?:[\s\-][\w\'\u00C0-\u024F]+)*?)'
                    r'\s+(' + INIT_SEP + r')\s*[,\.]',
                    tok, re.UNICODE
                )
                if m_start:
                    author_count += 1
                break  # either way, title starts after this

    if has_et_al or author_count >= 3:
        return first_lastname, '3+', has_et_al
    return first_lastname, author_count, False

# ── unified entry point ───────────────────────────────────────────────────────

def make_abbrev(citation):
    if not citation:
        return None
    c = str(citation).strip()
    year = extract_year(c)

    if is_format_b(c):
        lastname, count, et_al = parse_format_b(c)
    else:
        lastname, count, et_al = parse_format_a(c)

    if not lastname or not year:
        return None

    name = camel_case_name(lastname)

    if count == '3+' or et_al or (isinstance(count, int) and count >= 3):
        return f"{name}_{year}"
    elif count == 2:
        return f"{name}_{year}"   # we only have first author here; handle 2-author below
    else:
        return f"{name}_{year}"

# For 2-author case we need both lastnames — rewrite to return both
def parse_two_author(citation):
    """If exactly 2 authors (Format A), return (ln1, ln2). Else None."""
    c = re.sub(r'^\d+\.?\s*', '', str(citation).strip())
    if is_format_b(c):
        _, count, _ = parse_format_b(c)
        if count != 2:
            return None
        # Re-parse to get both names
        authors = []
        pos = 0
        while pos < len(c) and len(authors) < 2:
            m = re.match(r'([A-Z](?:\.[A-Z])?\.)\s+([A-Z][a-z]+)', c[pos:])
            if not m:
                break
            authors.append(m.group(2).split()[0])
            pos += m.end()
            sep = re.match(r',\s*', c[pos:])
            if sep:
                pos += sep.end()
            else:
                break
        return (authors[0], authors[1]) if len(authors) == 2 else None
    else:
        _, count, _ = parse_format_a(c)
        if count != 2:
            return None
        tokens = re.split(r',\s*', c)
        lns = []
        for i, tok in enumerate(tokens[:3]):
            tok = tok.strip()
            if re.search(r'\bet\.?\s*al\b', tok, re.IGNORECASE):
                break
            if i == 0:
                ln = _extract_lastname_from_start(tok)
            else:
                # Try end-of-token match first, then start-of-token
                INIT_SEP = r'[A-Z][A-Z\.\-]*(?=[\s,\.]|$)'
                m = re.search(r'^(.*?)\s+' + INIT_RE + r'\s*\.?\s*$', tok, re.UNICODE)
                if m:
                    ln = m.group(1).strip()
                else:
                    m2 = re.match(
                        r'^([\w\'\u00C0-\u024F]+(?:[\s\-][\w\'\u00C0-\u024F]+)*?)'
                        r'\s+(' + INIT_SEP + r')\s*[,\.]',
                        tok, re.UNICODE
                    )
                    ln = m2.group(1).strip() if m2 else None
            if ln:
                lns.append(ln)
            else:
                break
            if len(lns) == 2:
                break
        return (lns[0], lns[1]) if len(lns) == 2 else None

def make_abbrev_full(citation):
    """Compute abbrev_citation with proper 2-author handling."""
    if not citation:
        return None
    c = str(citation).strip()
    year = extract_year(c)
    if not year:
        return None

    if is_format_b(c):
        lastname, count, et_al = parse_format_b(c)
    else:
        lastname, count, et_al = parse_format_a(c)

    if not lastname:
        return None

    name1 = camel_case_name(lastname)

    if count == '3+' or et_al or (isinstance(count, int) and count >= 3):
        return f"{name1}_{year}"

    if isinstance(count, int) and count == 2:
        pair = parse_two_author(c)
        if pair:
            n1 = camel_case_name(pair[0])
            n2 = camel_case_name(pair[1])
            return f"{n1}&{n2}_{year}"
        return f"{name1}_{year}"

    # count == 1
    return f"{name1}_{year}"

# ── read, process, verify ─────────────────────────────────────────────────────

wb = load_workbook(INPUT)
all_rows = []

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        citing_paper, citing_doi, local_number, full_citation = row[0], row[1], row[2], row[3]
        if full_citation is None:
            continue
        abbrev = make_abbrev_full(full_citation)
        all_rows.append({
            'citing_paper': citing_paper,
            'citing_doi': citing_doi,
            'local_number': local_number,
            'full_citation': full_citation,
            'abbrev_citation': abbrev,
        })

print(f"Total references parsed: {len(all_rows)}")

# ── verification: 5 random rows ───────────────────────────────────────────────

random.seed(42)
sample = random.sample(all_rows, 5)
print("\n── 5 random rows (verification) ──")
for r in sample:
    c = str(r['full_citation'])
    fmt = "B" if is_format_b(c) else "A"
    if is_format_b(c):
        ln, count, et_al = parse_format_b(c)
    else:
        ln, count, et_al = parse_format_a(c)
    yr = extract_year(c)
    print(f"\nSheet: {r['citing_paper']}  |  local#: {r['local_number']}")
    print(f"Source: {c[:120]!r}")
    print(f"Format={fmt}  |  lastname={ln!r}  |  count={count}  |  et_al={et_al}  |  year={yr}")
    print(f"→ abbrev: {r['abbrev_citation']}")

# ── write output ──────────────────────────────────────────────────────────────

out_wb = Workbook()
ws_out = out_wb.active
ws_out.title = "parsed"
ws_out.append(['citing_paper', 'citing_doi', 'local_number', 'full_citation', 'abbrev_citation'])
for r in all_rows:
    ws_out.append([
        r['citing_paper'],
        r['citing_doi'],
        r['local_number'],
        r['full_citation'],
        r['abbrev_citation'],
    ])

out_wb.save(OUTPUT)
print(f"\nSaved → {OUTPUT}")
