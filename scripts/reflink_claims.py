#!/usr/bin/env python3
"""
Reference-link a claims workbook against a citing dictionary.

Applied to every sheet in the claims workbook EXCEPT sheets named "Quotes"
(case-insensitive) -- i.e. the Claims and Studies sheets are processed, the
Quotes sheet is ignored.

Two steps per sheet:

1. Expand the "Cited" column so there is exactly one reference per row. Each
   input row whose "Cited" cell holds multiple references separated by
   semicolons is duplicated once per reference. A purely-numeric chunk that
   uses commas (e.g. "74, 75") or a dash range (e.g. "111-114") is also split.
   Rows with an empty / missing "Cited" cell are kept as-is (Cited left blank).

2. Join each expanded row to the citing dictionary and append three columns:
       citing_id  <- dict "citing_id"
       ref_id     <- dict "ref_id"
       ref_label  <- dict "ref_bibtex"

   Two kinds of "Cited" references are supported:

   a. NUMERIC references (the common case). Matched on
          (claims "Citing" == dict "citing_bibtex") AND
          (claims "Cited"  == dict "local_number").
      Bracketed numbers such as "[57]" are accepted too.

   b. AUTHOR-YEAR in-text references (e.g. "Eriguchi et al. 2015",
      "Riwes and Reddy 2020", "Stein-Thoeringer 2022"). These occur when a
      citing paper cites by name rather than number (e.g. AzharUdDin_2025).
      The author surname and year are extracted by regex ("et al." is stripped,
      only the first author of an "X and Y" pair is used) and matched against
      the dict "ref_bibtex" values (formatted "Author_Year") belonging to the
      same citing paper. Matching is done on a normalised key with three tiers:
        1. exact author + year + disambiguation-letter suffix,
        2. author + year, ignoring the suffix (if the match is unique),
        3. author + year where the bibtex author ends with the cited surname,
           handling keys that prepend a first name (if the match is unique).

   When NO reference can be found and the "Cited" cell is not blank, the "Cited"
   cell in the output is highlighted RED for manual review.

Usage:
    python3 reflink_claims.py <claims.xlsx> <citing_dictionary.csv> <output.xlsx>
"""

import csv
import re
import sys

import openpyxl
from openpyxl.styles import PatternFill


NEW_COLUMNS = ["citing_id", "ref_id", "ref_label"]
IGNORE_SHEETS = {"quotes"}  # sheet names to skip (case-insensitive)
UNMATCHED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


# --------------------------------------------------------------------------- #
# Cited-cell expansion
# --------------------------------------------------------------------------- #
def parse_cited(value):
    """Expand a 'Cited' cell into a list of individual references.

    Returns [None] for empty/missing cells so the row is preserved once.
    Numeric items become ints; author-year items stay as strings.
    """
    if value is None:
        return [None]

    text = str(value).strip()
    if text == "":
        return [None]

    items = []
    for chunk in text.split(";"):
        chunk = chunk.strip()
        if chunk == "":
            continue
        # A purely-numeric chunk may use commas or dash ranges; split those.
        if "," in chunk and re.fullmatch(r"[\d\s,\-–—]+", chunk):
            subtokens = [t.strip() for t in chunk.split(",") if t.strip()]
        else:
            subtokens = [chunk]
        for token in subtokens:
            items.extend(_expand_token(token))
    return items if items else [None]


def _expand_token(token):
    """Expand a single token; dash ranges become multiple ints."""
    token = token.strip()
    # Strip surrounding brackets, e.g. "[57]" -> "57".
    token = re.sub(r"^\s*\[(.*)\]\s*$", r"\1", token).strip()
    if token == "":
        return [None]
    m = re.match(r"^(\d+)\s*[-–—]\s*(\d+)$", token)
    if m:
        start, end = int(m.group(1)), int(m.group(2))
        step = 1 if end >= start else -1
        return list(range(start, end + step, step))
    if re.match(r"^\d+$", token):
        return [int(token)]
    return [token]


# --------------------------------------------------------------------------- #
# Dictionary loading
# --------------------------------------------------------------------------- #
def _clean(v):
    if v is None:
        return ""
    v = v.strip()
    return "" if v.upper() == "NA" else v


def _norm(s):
    """Lowercase and strip everything but letters/digits."""
    return re.sub(r"[^a-z0-9]", "", s.lower())


def load_dictionary(dict_path):
    """Return (by_local, by_author).

    by_local : {(citing_bibtex, local_number): payload}
    by_author: {citing_bibtex: [ (author_norm, year, suffix, payload), ... ]}
    where payload = (citing_id, ref_id, ref_bibtex).
    """
    by_local = {}
    by_author = {}
    with open(dict_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        required = {"citing_bibtex", "local_number", "citing_id", "ref_id", "ref_bibtex"}
        missing = required - set(reader.fieldnames or [])
        if missing:
            sys.exit(f"Dictionary is missing required columns: {sorted(missing)}")

        for row in reader:
            bibtex = _clean(row["citing_bibtex"])
            if not bibtex:
                continue
            payload = (_clean(row["citing_id"]), _clean(row["ref_id"]), _clean(row["ref_bibtex"]))

            # Numeric index (local_number).
            local = _clean(row["local_number"])
            if local:
                for key in {(bibtex, local), _int_key(bibtex, local)}:
                    if key is not None:
                        by_local.setdefault(key, payload)

            # Author-year index (parsed from ref_bibtex "Author_Year").
            parsed = _parse_ref_bibtex(payload[2])
            if parsed is not None:
                author_norm, year, suffix = parsed
                by_author.setdefault(bibtex, []).append((author_norm, year, suffix, payload))

    return by_local, by_author


def _int_key(bibtex, local):
    try:
        return (bibtex, int(float(local)))
    except ValueError:
        return None


def _parse_ref_bibtex(ref_bibtex):
    """Parse 'Author_Year[suffix]' -> (author_norm, year, suffix) or None."""
    if not ref_bibtex or "_" not in ref_bibtex:
        return None
    author, year = ref_bibtex.rsplit("_", 1)
    m = re.match(r"^(\d{4})([a-z]*)$", year)
    if not m:
        return None
    author_norm = _norm(author)
    if not author_norm:
        return None
    return author_norm, m.group(1), m.group(2)


# --------------------------------------------------------------------------- #
# Lookup
# --------------------------------------------------------------------------- #
def lookup_numeric(by_local, citing, cited):
    if citing is None or cited is None:
        return None
    citing = str(citing).strip()
    for key in ((citing, cited), (citing, str(cited).strip())):
        if key in by_local:
            return by_local[key]
    try:
        return by_local.get((citing, int(float(str(cited).strip()))))
    except ValueError:
        return None


def _parse_intext(token):
    """Parse an author-year in-text citation -> (author_norm, year, suffix)."""
    t = re.sub(r"\bet\s+al\.?", " ", str(token), flags=re.IGNORECASE)
    m = re.search(r"(\d{4})([a-z]?)", t)
    if not m:
        return None
    year, suffix = m.group(1), m.group(2).lower()
    author = t[: m.start()]
    # First author only for "X and Y" / "X & Y" / "X, Y".
    author = re.split(r"\s+and\s+|\s*&\s*|,", author)[0]
    author_norm = _norm(author)
    if not author_norm:
        return None
    return author_norm, year, suffix


def lookup_author(by_author, citing, token):
    if citing is None or token is None:
        return None
    parsed = _parse_intext(token)
    if parsed is None:
        return None
    author_norm, year, suffix = parsed
    entries = by_author.get(str(citing).strip(), [])

    # Tier 1: exact author + year + suffix.
    exact = [e for e in entries if e[0] == author_norm and e[1] == year and e[2] == suffix]
    if exact:
        return exact[0][3]
    # Tier 2: author + year, ignore suffix (unique only).
    yr = [e for e in entries if e[0] == author_norm and e[1] == year]
    if len(yr) == 1:
        return yr[0][3]
    # Tier 3: bibtex author ends with cited surname (prepended first name), unique.
    ew = [e for e in entries if e[1] == year and e[0].endswith(author_norm)]
    if len(ew) == 1:
        return ew[0][3]
    return None


# --------------------------------------------------------------------------- #
# Sheet processing
# --------------------------------------------------------------------------- #
def process_sheet(ws_in, ws_out, by_local, by_author, unmatched):
    header = [c.value for c in next(ws_in.iter_rows(min_row=1, max_row=1))]
    try:
        citing_idx = header.index("Citing")
        cited_idx = header.index("Cited")
    except ValueError:
        print(f"  ! Sheet '{ws_in.title}' has no Citing/Cited columns; skipped.")
        return None

    ws_out.append(list(header) + NEW_COLUMNS)

    rows_in = rows_out = 0
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        rows_in += 1
        row = list(row)
        citing = row[citing_idx]
        for ref in parse_cited(row[cited_idx]):
            new_row = list(row)
            new_row[cited_idx] = ref
            match = None
            if isinstance(ref, int):
                match = lookup_numeric(by_local, citing, ref)
            elif isinstance(ref, str):
                match = lookup_author(by_author, citing, ref)

            if match is None:
                new_row.extend(["", "", ""])
            else:
                new_row.extend(list(match))
            ws_out.append(new_row)
            rows_out += 1

            # Highlight the Cited cell red when unmatched and not blank.
            if match is None and ref is not None:
                ws_out.cell(row=ws_out.max_row, column=cited_idx + 1).fill = UNMATCHED_FILL
                unmatched.append((ws_in.title, citing, ref))
    return rows_in, rows_out


def main():
    if len(sys.argv) != 4:
        sys.exit("Usage: python3 reflink_claims.py <claims.xlsx> <citing_dictionary.csv> <output.xlsx>")
    claims_path, dict_path, output_path = sys.argv[1:4]

    by_local, by_author = load_dictionary(dict_path)
    wb_in = openpyxl.load_workbook(claims_path)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    unmatched = []
    for sheet_name in wb_in.sheetnames:
        if sheet_name.strip().lower() in IGNORE_SHEETS:
            print(f"[{sheet_name}] ignored")
            continue
        ws_in = wb_in[sheet_name]
        ws_out = wb_out.create_sheet(title=sheet_name)
        result = process_sheet(ws_in, ws_out, by_local, by_author, unmatched)
        if result is None:
            wb_out.remove(ws_out)
            continue
        rows_in, rows_out = result
        print(f"[{sheet_name}] {rows_in} input rows -> {rows_out} expanded rows")

    wb_out.save(output_path)
    print(f"Wrote {output_path}")

    if unmatched:
        print(f"\n{len(unmatched)} reference(s) had no dictionary match (highlighted red):")
        for sheet_name, citing, cited in unmatched:
            print(f"  [{sheet_name}] {citing} / {cited}")


if __name__ == "__main__":
    main()
