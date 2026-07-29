#!/usr/bin/env python3
"""
Expand bracketed reference lists in the last column of an Excel sheet.

Each input row whose final column contains a bracketed list such as
[29,105] or [105,106,111-114] is duplicated once per reference number,
with the final column replaced by a single integer.

Tokens may be:
  - a single number:        29        -> 29
  - a comma-separated list: 2,3       -> 2, 3
  - a dash range:           3-5       -> 3, 4, 5
Combinations work too: [2,3-5] -> 2, 3, 4, 5

Usage:
    python3 expand_brackets.py <input.xlsx> <output.xlsx>
"""

import re
import sys

import openpyxl


def parse_brackets(value):
    """Return a list of integers expanded from a bracketed reference string."""
    if value is None:
        return [None]

    text = str(value).strip()
    # Strip surrounding square brackets if present.
    text = re.sub(r"^\s*\[(.*)\]\s*$", r"\1", text)
    text = text.strip()
    if text == "":
        return [None]

    numbers = []
    for token in text.split(","):
        token = token.strip()
        if token == "":
            continue
        # Range like 111-114 (also handles unicode en/em dashes).
        m = re.match(r"^(\d+)\s*[-–—]\s*(\d+)$", token)
        if m:
            start, end = int(m.group(1)), int(m.group(2))
            step = 1 if end >= start else -1
            numbers.extend(range(start, end + step, step))
        elif re.match(r"^\d+$", token):
            numbers.append(int(token))
        else:
            # Not a pure number/range; keep the original token as-is.
            numbers.append(token)

    return numbers if numbers else [None]


def expand_sheet(input_path, output_path):
    wb_in = openpyxl.load_workbook(input_path)
    ws_in = wb_in.active

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = ws_in.title

    for row in ws_in.iter_rows(values_only=True):
        if row is None or len(row) == 0:
            continue
        last = row[-1]
        for num in parse_brackets(last):
            ws_out.append(list(row[:-1]) + [num])

    wb_out.save(output_path)
    print(f"Wrote {output_path} ({ws_out.max_row} rows from {ws_in.max_row} input rows)")


def main():
    if len(sys.argv) != 3:
        sys.exit("Usage: python3 expand_brackets.py <input.xlsx> <output.xlsx>")
    expand_sheet(sys.argv[1], sys.argv[2])


if __name__ == "__main__":
    main()
