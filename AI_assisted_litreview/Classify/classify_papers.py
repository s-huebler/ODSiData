import openpyxl
import csv
import re
import os

INPUT = "Classify/Included.xlsx"
OUTPUT_DIR = "Classify"

STUDY_TYPES = [
    "CaseStudy", "RCT", "CaseControl", "MetaAnalysis",
    "Cohort", "Review", "Multiple", "Other"
]

# Keywords for each type — matched against title+abstract (case-insensitive)
PATTERNS = {
    "CaseStudy": [
        r"\bcase[ -]stud", r"\bcase[ -]series", r"\bcase report"
    ],
    "RCT": [
        r"\brandomized controlled trial", r"\brandomised controlled trial",
        r"\bRCT\b", r"\brandomly assigned", r"\brandom assignment",
        r"\bdouble[ -]blind", r"\bplacebo[ -]controlled",
        r"\bphase (II|III|2|3) (trial|study)",
    ],
    "CaseControl": [
        r"\bcase[ -]control", r"\bmatched control", r"\bcontrols? were matched",
        r"\bcases? and controls?\b", r"\bcases? versus controls?\b",
        r"\bcases? vs\.? controls?\b",
        r"\bcompared (with|to) (matched |source-matched )?(control patients|controls)\b",
    ],
    "MetaAnalysis": [
        r"\bmeta[ -]anal", r"\bsystematic review and meta",
        r"\bpooled anal", r"\bdata synthesis", r"\bpooled estimate"
    ],
    "Cohort": [
        r"\bcohort", r"\bprospective stud", r"\bretrospective stud",
        r"\bcross[ -]sectional", r"\blongitudinal stud",
        r"\bobservational stud",
        # common SCT study language without "cohort"
        r"\bin a retrospective", r"\bwe retrospectively",
        r"\bwe prospectively", r"\bwe analyzed\b", r"\bwe evaluated\b",
        r"\bwe enrolled\b", r"\bwere enrolled\b", r"\bpatients were enrolled",
        r"\bwe included\b", r"\bwe recruited\b",
        r"\bsingle[ -]center stud", r"\bmulti[ -]center stud",
        r"\bmulticenter stud", r"\bsingle[ -]institution",
        r"\bpatients who (under|receiv)", r"\bpatients undergoing",
        r"\bstool samples were collected", r"\bfecal samples were collected",
        r"\bwe collected\b.*sample", r"\bsample collection",
    ],
    "Review": [
        r"\bsystematic review\b", r"\bscoping review\b",
        r"\bnarrative review\b", r"\bliterature review\b",
        r"\breview article\b", r"\breview of the literature",
        r"\bwe review\b", r"\bthis review\b",
        r"\bin this review\b", r"\bthis article reviews",
        r"\bhere we review", r"\baims? to review",
        r"\bprovide[sd]? an overview", r"\bwe summarize\b",
        r"\bthis paper (reviews|summarizes|discusses|provides)",
        r"\bwe discuss\b.*overview", r"\boverview of",
    ],
}

def classify(title, abstract):
    if not abstract or str(abstract).strip() in ("", "None"):
        return "Other"

    text = f"{title or ''} {abstract or ''}".lower()

    hits = set()
    for stype, pats in PATTERNS.items():
        for pat in pats:
            if re.search(pat, text, re.IGNORECASE):
                hits.add(stype)
                break

    # MetaAnalysis supersedes Review and Cohort (meta-analyses contain cohort-like language)
    if "MetaAnalysis" in hits:
        hits.discard("Review")
        hits.discard("Cohort")

    # RCT supersedes Cohort and CaseControl (trials use those patterns incidentally)
    if "RCT" in hits:
        hits.discard("Cohort")
        hits.discard("CaseControl")

    # CaseControl supersedes Cohort (cohort language appears in case-control methods)
    if "CaseControl" in hits:
        hits.discard("Cohort")

    # Review supersedes Cohort only when no primary-study type is also present
    if "Review" in hits and hits == {"Review", "Cohort"}:
        hits.discard("Cohort")

    if len(hits) == 0:
        return "Other"
    if len(hits) == 1:
        return hits.pop()
    return "Multiple"


def main():
    wb = openpyxl.load_workbook(INPUT)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    title_idx = headers.index("Title")
    abstract_idx = headers.index("Abstract")

    # Collect rows by study type
    buckets = {t: [] for t in STUDY_TYPES}

    for row in ws.iter_rows(min_row=2, values_only=True):
        title = row[title_idx]
        abstract = row[abstract_idx]
        stype = classify(title, abstract)
        buckets[stype].append(row)

    # Write one CSV per non-empty bucket
    for stype, rows in buckets.items():
        if not rows:
            continue
        out_path = os.path.join(OUTPUT_DIR, f"{stype}_classified.csv")
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)
        print(f"{stype}: {len(rows)} papers → {out_path}")

    total = sum(len(r) for r in buckets.values())
    print(f"\nTotal classified: {total} (input rows: {ws.max_row - 1})")


if __name__ == "__main__":
    main()
