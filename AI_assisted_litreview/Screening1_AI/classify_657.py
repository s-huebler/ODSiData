"""
First-pass title/abstract screening for ODSiData GVHD microbiome meta-analysis.
657 papers from InitialList657_Title.csv.
"""

import csv
import re
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font

INPUT  = "Screening1_AI/InitialList657_Title.csv"
OUTPUT = "Screening1_AI/InitialList657_classified.xlsx"


def norm(text):
    """Lowercase + collapse whitespace for matching."""
    return re.sub(r"\s+", " ", (text or "").lower())


# ── COHORT helpers ─────────────────────────────────────────────────────────────

ALLO_SIGNALS = [
    "allogeneic", "allogenic", "allo-hct", "allo-sct", "allo-hsct", "allo-bmt",
    "allo hct", "allo sct", "allo hsct", "allo bmt",
    "allograft", "allogeneic transplant", "allogeneic hematopoietic",
    "allogeneic stem cell", "allogeneic bone marrow",
    "matched related donor", "matched unrelated donor", "haploidentical",
    "cord blood transplant", "umbilical cord blood transplant",
    "donor stem cell", "stem cell donor",
]

def has_allo_cohort(t):
    return any(s in t for s in ALLO_SIGNALS)

def is_auto_only(t):
    """True only if abstract/title mentions autologous but NOT allogeneic."""
    has_auto = "autologous" in t
    has_allo = has_allo_cohort(t)
    return has_auto and not has_allo


# ── GUT MICROBIOME helpers ─────────────────────────────────────────────────────

# Explicit gut microbiome terms (each alone is sufficient)
GUT_EXPLICIT = [
    "gut microbiome", "gut microbiota", "gut microbi",
    "intestinal microbiome", "intestinal microbiota", "intestinal microbi",
    "fecal microbiota", "fecal microbiome", "faecal microbiota", "faecal microbiome",
    "stool microbiome", "stool microbiota",
    "gut dysbiosis", "intestinal dysbiosis",
    "fecal microbiota transplant", "faecal microbiota transplant",
    "fecal transplant", "faecal transplant",
    "gut bacteria", "intestinal bacteria",
    "gut flora", "intestinal flora", "gut microflora", "intestinal microflora",
    "intestinal microbial", "gut microbial",
    "gut colonization", "intestinal colonization",
    "intestinal commensals", "gut commensals",
]

# Named taxa that in allo-HSCT papers almost always appear in gut microbiome studies
GUT_TAXA = [
    "blautia", "faecalibacterium", "akkermansia", "lachnospiraceae",
    "ruminococcaceae", "clostridiales", "bacteroidetes", "fusobacterium",
    "eubacterium limosum", "lachnospira", "roseburia",
]

# Fecal sample collection signals
FECAL_SAMPLE = [
    "stool sample", "fecal sample", "faecal sample",
    "stool collection", "fecal collection",
    "stool culture", "fecal culture",
]

# Metabolites that in allo-HSCT context imply gut microbiome work
GUT_METABOLITES_W_MICRO = [
    "butyrate", "short-chain fatty acid", " scfa", "(scfa)",
    "secondary bile acid", "microbial metabolite",
]

# Sequencing signals that are gut-specific when paired with gut context
SEQ_SIGNALS = ["16s rrna", "16s sequenc", "shotgun metagen", "metagenom"]

# Probiotic signals that imply gut microbiome work
PROBIOTIC_SIGNALS = ["lactobacillus", "bifidobacterium", "probiotic", "prebiotic", " fmt ", "(fmt)"]

# Non-gut site terms — if ONLY these appear with "microbi" and no gut terms → not qualifying
NON_GUT_ONLY = ["oral microbiome", "oral microbiota", "skin microbiome", "skin microbiota",
                "cutaneous microbiome", "cutaneous microbiota", "vaginal microbiome",
                "lung microbiome", "pulmonary microbiome", "ocular surface microbiota",
                "airway microbiome", "nasal microbiome"]


def has_gut_microbiome(t):
    """Return (bool, signal_description)."""

    # 1. Explicit gut microbiome terms
    for s in GUT_EXPLICIT:
        if s in t:
            return True, f"explicit gut microbiome term: '{s}'"

    # 2. Named gut taxa
    for s in GUT_TAXA:
        if s in t:
            return True, f"named gut taxon: '{s}'"

    # 3. Fecal sample collection
    for s in FECAL_SAMPLE:
        if s in t:
            return True, f"fecal sample collection: '{s}'"

    # 4. Gut metabolites + any microbiome mention
    gut_context = any(g in t for g in ["gut", "intestin", "fecal", "faecal", "stool", "microbi", "microflora"])
    for s in GUT_METABOLITES_W_MICRO:
        if s in t and gut_context:
            return True, f"gut metabolite in microbiome context: '{s}'"

    # 5. Sequencing signals + gut context
    for s in SEQ_SIGNALS:
        if s in t and any(g in t for g in ["gut", "intestin", "fecal", "faecal", "stool"]):
            return True, f"sequencing ({s}) with gut context"

    # 6. Probiotic/FMT signals + microbiome mention
    for s in PROBIOTIC_SIGNALS:
        if s in t and ("microbi" in t or "microflora" in t or "intestin" in t or "gut" in t):
            return True, f"probiotic/FMT signal: '{s}' with microbiome context"

    # 7. "microbial diversity" / "microbiome diversity" with gut context
    if ("microbial diversity" in t or "microbiome diversity" in t or "microbiota diversity" in t):
        if any(g in t for g in ["gut", "intestin", "fecal", "faecal", "stool"]):
            return True, "microbial diversity with gut context"

    # 8. Generic "microbiome" / "microbiota" in non-gut-only context
    has_microbi = "microbiome" in t or "microbiota" in t or "microbi" in t
    is_non_gut_only = any(s in t for s in NON_GUT_ONLY) and not any(
        g in t for g in ["gut microbi", "intestinal microbi", "fecal microbi", "stool microbi"]
    )
    if has_microbi and not is_non_gut_only:
        # Generic mention — flag as borderline (lean-Include)
        return True, "generic microbiome/microbiota mention (lean-Include applied)"

    return False, ""


# ── OUTCOME helpers ────────────────────────────────────────────────────────────

GVHD_SIGNALS = [
    "gvhd", "graft-versus-host", "graft versus host", "graft vs host",
    "agvhd", "cgvhd", "gvhd",
]

BROADER_OUTCOME_SIGNALS = [
    "overall survival", "non-relapse mortality", "nonrelapse mortality",
    "transplant-related mortality", "transplant related mortality",
    "trm", "nrm", "engraftment", "relapse", "relapse-free survival",
    "infection", "bacteremia", "bloodstream infection",
    "post-transplant", "posttransplant", "post transplant",
    "transplant outcome", "transplant complication",
    "adverse event", "adverse effect",
    "immunologic complication", "morbidity", "mortality",
]


def has_gvhd(t):
    return any(s in t for s in GVHD_SIGNALS)

def has_broader_outcome(t):
    return any(s in t for s in BROADER_OUTCOME_SIGNALS)


# ── Main classifier ─────────────────────────────────────────────────────────────

def classify(title, abstract):
    t = norm(title + " " + abstract)
    t_title = norm(title)

    # ── Cohort check
    if is_auto_only(t):
        return "Discard", (
            "Cohort dimension not satisfied: autologous-only transplantation context; "
            "no allogeneic transplant."
        )

    if not has_allo_cohort(t):
        return "Discard", (
            "Cohort dimension not satisfied: no allogeneic stem cell / hematopoietic cell / "
            "bone marrow transplantation context detected."
        )

    # ── Gut microbiome check
    gut_ok, gut_signal = has_gut_microbiome(t)
    if not gut_ok:
        return "Discard", (
            "Cohort (allo-SCT) satisfied, but gut microbiome dimension not satisfied: "
            "no qualifying gut microbiome signals detected."
        )

    # ── Outcome check
    gvhd_ok = has_gvhd(t)
    broader_ok = has_broader_outcome(t)

    if gvhd_ok:
        return "Include", (
            f"All three dimensions satisfied: allo-SCT cohort, gut microbiome ({gut_signal}), "
            f"and GVHD as outcome."
        )

    if broader_ok:
        # Accept broader outcomes only when gut microbiome is substantive focus
        is_substantive = "generic" not in gut_signal
        if is_substantive:
            return "Include", (
                f"All three dimensions satisfied: allo-SCT cohort, gut microbiome as substantive "
                f"focus ({gut_signal}), and broader allo-SCT outcome."
            )
        else:
            return "Discard", (
                "Gut microbiome only generically mentioned; broader allo-SCT outcome present but "
                "microbiome is not a substantive focus — insufficient for Include without GVHD."
            )

    # No outcome detected — discard
    return "Discard", (
        "Cohort (allo-SCT) and gut microbiome satisfied, but outcome dimension not satisfied: "
        "no GVHD or qualifying allo-SCT outcome detected."
    )


# ── Run ─────────────────────────────────────────────────────────────────────────

rows = []
with open(INPUT, newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for row in reader:
        rows.append(row)

results = []
for row in rows:
    decision, reasoning = classify(row.get("Title", ""), row.get("Abstract", ""))
    results.append((row["Title"], row["Abstract"], decision, reasoning))

# ── Write xlsx ───────────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Classified"

headers = ["Title", "Abstract", "Decision", "Reasoning"]
ws.append(headers)

# Header style
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = h
    cell.fill = header_fill
    cell.font = header_font

include_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
discard_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

for i, (title, abstract, decision, reasoning) in enumerate(results, start=2):
    ws.cell(row=i, column=1, value=title)
    ws.cell(row=i, column=2, value=abstract)
    ws.cell(row=i, column=3, value=decision)
    ws.cell(row=i, column=4, value=reasoning)
    fill = include_fill if decision == "Include" else discard_fill
    for col in range(1, 5):
        c = ws.cell(row=i, column=col)
        c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="top")

ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 80
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 70

# Freeze header row
ws.freeze_panes = "A2"

wb.save(OUTPUT)

include_n = sum(1 for _, _, d, _ in results if d == "Include")
discard_n = sum(1 for _, _, d, _ in results if d == "Discard")
print(f"Saved: {OUTPUT}")
print(f"Include: {include_n}")
print(f"Discard: {discard_n}")
print(f"Total:   {include_n + discard_n}")

# Quick breakdown of discard reasons
from collections import Counter
discard_reasons = []
for _, _, d, r in results:
    if d == "Discard":
        if "Cohort dimension not satisfied: no allogeneic" in r:
            discard_reasons.append("No allo-SCT cohort")
        elif "autologous-only" in r:
            discard_reasons.append("Autologous-only")
        elif "gut microbiome dimension not satisfied" in r:
            discard_reasons.append("No gut microbiome")
        elif "generic" in r.lower():
            discard_reasons.append("Generic microbiome + no GVHD")
        elif "outcome dimension not satisfied" in r:
            discard_reasons.append("No qualifying outcome")
        else:
            discard_reasons.append("Other")

c = Counter(discard_reasons)
print("\nDiscard breakdown:")
for reason, count in c.most_common():
    print(f"  {reason}: {count}")
